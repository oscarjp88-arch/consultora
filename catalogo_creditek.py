#!/usr/bin/env python3
"""
CREDITEK · Generador de catálogo enriquecido v2
- Specs vía Gemini API (reemplaza scraping GSMArena)
- Modo incremental: solo procesa referencias nuevas
- Sync automático: genera catalogo.json + commit + push al final
"""

import os
import re
import json
import time
import io
import sys
import subprocess
from datetime import datetime

try:
    import requests
    import openpyxl
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
except ImportError:
    print("⚠ Faltan librerías. Instalando...")
    os.system(f"{sys.executable} -m pip install --quiet --break-system-packages "
              "requests openpyxl google-auth-oauthlib google-api-python-client")
    print("✓ Librerías instaladas. Vuelve a ejecutar el script.")
    sys.exit(0)

# === RUTAS ===
REPO_PATH     = '/Users/oscarpacheco/consultora'
CREDS_PATH    = f'{REPO_PATH}/credentials.json'
TOKEN_PATH    = f'{REPO_PATH}/token.json'
ENV_PATH      = f'{REPO_PATH}/.env'
DATA_DIR      = f'{REPO_PATH}/creditek/data'
SPECS_DIR     = f'{DATA_DIR}/specs'
IMGS_DIR      = f'{REPO_PATH}/creditek/assets/imagenes'

# === IDs DRIVE ===
EXCEL_FILE_ID = '1givs_1TU6ug3QrrdhgYXm6Z15kkdemin'
CARPETA_BASE  = '1QtZ4K1770F3uXrbic9_wPZtfgy6U-58y'
SCOPES        = ['https://www.googleapis.com/auth/drive']

MARCAS_RECONOCIDAS = {
    'SAMSUNG', 'APPLE', 'IPHONE', 'XIAOMI', 'MOTOROLA', 'MOTO', 'HONOR', 'OPPO',
    'REALME', 'INFINIX', 'TECNO', 'NOKIA', 'ZTE', 'HUAWEI', 'VIVO', 'ALCATEL',
    'POCO', 'TCL', 'IFFALCON', 'ITEL', 'JBL', 'BOSE', 'ACER', 'HP', 'ASUS',
    'LENOVO', 'IPAD', 'NUBIA'
}
MARCAS_SOLO_SPECS = {'CORN', 'KRONO', 'FLY', 'HYUNDAI', 'NET', 'XKIM', 'TABLET', 'PORTÁTIL', 'PORTATIL'}

GEMINI_MODEL  = 'gemini-2.5-flash'
GEMINI_DELAY  = 5   # segundos entre llamadas (free tier: 15 RPM)

# === ENV ===
def cargar_env():
    env = {}
    if os.path.exists(ENV_PATH):
        with open(ENV_PATH) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    k, v = line.split('=', 1)
                    env[k.strip()] = v.strip()
    return env

# === AUTH DRIVE ===
def autenticar_drive():
    creds = None
    if os.path.exists(TOKEN_PATH):
        creds = Credentials.from_authorized_user_file(TOKEN_PATH, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDS_PATH, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_PATH, 'w') as f:
            f.write(creds.to_json())
    return build('drive', 'v3', credentials=creds)

# === DRIVE HELPERS ===
def buscar_o_crear_carpeta(svc, nombre, parent_id):
    q = f"name='{nombre}' and '{parent_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
    res = svc.files().list(q=q, fields='files(id,name)').execute().get('files', [])
    if res:
        return res[0]['id']
    meta = {'name': nombre, 'mimeType': 'application/vnd.google-apps.folder', 'parents': [parent_id]}
    return svc.files().create(body=meta, fields='id').execute()['id']

def archivo_existe(svc, nombre, parent_id):
    q = f"name='{nombre}' and '{parent_id}' in parents and trashed=false"
    res = svc.files().list(q=q, fields='files(id)').execute().get('files', [])
    return res[0]['id'] if res else None

def subir_archivo(svc, nombre, contenido_bytes, mimetype, parent_id):
    existente = archivo_existe(svc, nombre, parent_id)
    media = MediaIoBaseUpload(io.BytesIO(contenido_bytes), mimetype=mimetype, resumable=False)
    if existente:
        return svc.files().update(fileId=existente, media_body=media).execute()['id']
    meta = {'name': nombre, 'parents': [parent_id]}
    return svc.files().create(body=meta, media_body=media, fields='id').execute()['id']

def descargar_archivo_drive(svc, file_id):
    req = svc.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return buf

def listar_archivos_carpeta(svc, carpeta_id):
    """Devuelve lista de {name, id} en una carpeta de Drive."""
    resultados = []
    token = None
    while True:
        res = svc.files().list(
            q=f"'{carpeta_id}' in parents and trashed=false",
            fields='nextPageToken, files(id, name)',
            pageToken=token
        ).execute()
        resultados.extend(res.get('files', []))
        token = res.get('nextPageToken')
        if not token:
            break
    return resultados

# === UTILIDADES ===
def slug(texto):
    s = re.sub(r'[^\w\s.-]', '', str(texto), flags=re.UNICODE)
    s = re.sub(r'\s+', '_', s.strip())
    return s.upper()[:90]

def detectar_marca(referencia):
    ref_upper = str(referencia).upper()
    palabras = re.findall(r'[A-ZÁÉÍÓÚÑ]+', ref_upper)
    for p in palabras:
        if p in MARCAS_RECONOCIDAS:
            return {'MOTO': 'MOTOROLA', 'IPHONE': 'APPLE', 'IPAD': 'APPLE', 'NUBIA': 'ZTE'}.get(p, p)
    for p in palabras:
        if p in MARCAS_SOLO_SPECS:
            return p
    return palabras[0] if palabras else 'GENERICO'

def es_marca_reconocida(marca):
    return marca.upper() in MARCAS_RECONOCIDAS

# === GEMINI API ===
def consultar_gemini(referencia, api_key):
    """Devuelve dict con specs e imagen_url, o None en caso de error."""
    prompt = (
        f'Para el producto "{referencia}", responde SOLO con JSON válido '
        '(sin markdown, sin texto adicional):\n'
        '{"pantalla":"","ram_almacenamiento":"","camara":"","bateria":"",'
        '"red":"","sistema":"","imagen_url":null,"confianza":""}\n\n'
        'Reglas:\n'
        '- imagen_url: URL directa a imagen JPG del fabricante oficial '
        '(samsung.com, hihonor.com, motorola.com, etc.) o null si no estás seguro\n'
        '- confianza: alta (modelo exacto conocido) / media / baja\n'
        '- Campos vacíos: string vacío "", nunca null (excepto imagen_url)\n'
        '- Solo el JSON, nada más'
    )
    url = (f'https://generativelanguage.googleapis.com/v1beta/models/'
           f'{GEMINI_MODEL}:generateContent?key={api_key}')
    payload = {
        'contents': [{'parts': [{'text': prompt}]}],
        'generationConfig': {'maxOutputTokens': 2000, 'temperature': 0}
    }
    try:
        r = requests.post(url, json=payload, timeout=30)
        if r.status_code != 200:
            return None
        d = r.json()
        text = d['candidates'][0]['content']['parts'][0]['text']
        m = re.search(r'\{.*\}', text, re.DOTALL)
        if not m:
            return None
        return json.loads(m.group())
    except Exception:
        return None

# === IMAGEN ===
def descargar_imagen(url):
    if not url:
        return None
    try:
        H = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'}
        r = requests.get(url, headers=H, timeout=15)
        ct = r.headers.get('Content-Type', '')
        if r.status_code == 200 and 'image' in ct and len(r.content) > 5000:
            return r.content
    except Exception:
        pass
    return None

def buscar_imagen_wikipedia(referencia):
    """Fallback: busca imagen en Wikipedia REST API."""
    # Construir posible título de Wikipedia
    titulo = re.sub(r'\s+\d+/\d+GB.*$', '', referencia.strip(), flags=re.IGNORECASE)
    titulo = re.sub(r'\s+', '_', titulo.strip())
    try:
        H = {'User-Agent': 'creditek-catalog/1.0'}
        url = f'https://en.wikipedia.org/api/rest_v1/page/summary/{titulo}'
        r = requests.get(url, headers=H, timeout=8)
        if r.status_code == 200:
            img_url = r.json().get('thumbnail', {}).get('source')
            if img_url:
                return descargar_imagen(img_url)
    except Exception:
        pass
    return None

# === INDICE (MODO INCREMENTAL) ===
def cargar_indice_existente(svc):
    """Descarga INDICE.xlsx de Drive. Devuelve dict slug→row, o {} si no existe."""
    file_id = archivo_existe(svc, 'INDICE.xlsx', CARPETA_BASE)
    if not file_id:
        return {}
    buf = descargar_archivo_drive(svc, file_id)
    wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    ws = wb.active
    indice = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            indice[slug(str(row[0]))] = {
                'referencia': row[0],
                'marca': row[1] or '',
                'tiene_imagen': str(row[2] or '') == '✓',
                'tiene_specs': str(row[3] or '') == '✓',
                'fuente': row[4] or '',
                'fecha': row[5] or '',
                'notas': row[6] or '',
            }
    return indice

# === SYNC GITHUB PAGES ===
def sincronizar_github_pages(svc, id_specs, id_imgs, indice_actual):
    """Descarga archivos de Drive, genera catalogo.json, hace commit + push."""
    print("\n" + "="*70)
    print("  SYNC → GITHUB PAGES")
    print("="*70)

    os.makedirs(SPECS_DIR, exist_ok=True)
    os.makedirs(IMGS_DIR, exist_ok=True)
    os.makedirs(DATA_DIR, exist_ok=True)

    # 1. Descargar specs JSONs de Drive
    print("→ Descargando specs de Drive...")
    archivos_specs = listar_archivos_carpeta(svc, id_specs)
    for f in archivos_specs:
        try:
            buf = descargar_archivo_drive(svc, f['id'])
            local_path = os.path.join(SPECS_DIR, f['name'])
            with open(local_path, 'wb') as out:
                out.write(buf.read())
        except Exception as e:
            print(f"  ⚠ Error descargando {f['name']}: {e}")
    print(f"  ✓ {len(archivos_specs)} specs descargadas")

    # 2. Descargar imágenes de Drive (solo las que no existen localmente)
    print("→ Descargando imágenes de Drive...")
    archivos_imgs = listar_archivos_carpeta(svc, id_imgs)
    nuevas_imgs = 0
    for f in archivos_imgs:
        local_path = os.path.join(IMGS_DIR, f['name'])
        if not os.path.exists(local_path):
            try:
                buf = descargar_archivo_drive(svc, f['id'])
                with open(local_path, 'wb') as out:
                    out.write(buf.read())
                nuevas_imgs += 1
            except Exception as e:
                print(f"  ⚠ Error descargando {f['name']}: {e}")
    print(f"  ✓ {len(archivos_imgs)} imágenes en Drive, {nuevas_imgs} nuevas descargadas")

    # 3. Generar catalogo.json desde specs locales
    print("→ Generando catalogo.json...")
    catalogo = {}
    for fname in os.listdir(SPECS_DIR):
        if not fname.endswith('.json'):
            continue
        slug_key = fname[:-5]  # quitar .json
        try:
            with open(os.path.join(SPECS_DIR, fname), encoding='utf-8') as f:
                data = json.load(f)
            img_path = os.path.join(IMGS_DIR, slug_key + '.jpg')
            catalogo[slug_key] = {
                'pantalla':          data.get('pantalla', ''),
                'ram_almacenamiento': data.get('ram_almacenamiento', ''),
                'camara':            data.get('camara', ''),
                'bateria':           data.get('bateria', ''),
                'red':               data.get('red', ''),
                'sistema':           data.get('sistema', ''),
                'tiene_imagen':      os.path.exists(img_path),
                'confianza':         data.get('confianza', ''),
            }
        except Exception:
            pass

    catalogo_path = os.path.join(DATA_DIR, 'catalogo.json')
    with open(catalogo_path, 'w', encoding='utf-8') as f:
        json.dump(catalogo, f, ensure_ascii=False, indent=2)
    print(f"  ✓ catalogo.json con {len(catalogo)} referencias")

    # 4. Git commit + push
    con_img = sum(1 for v in catalogo.values() if v['tiene_imagen'])
    fecha = datetime.now().strftime('%Y-%m-%d')
    msg = f"catálogo: sync {fecha} ({len(catalogo)} refs, {con_img} con imagen)"

    try:
        subprocess.run(['git', '-C', REPO_PATH, 'add',
                        'creditek/data/', 'creditek/assets/imagenes/'],
                       check=True, capture_output=True)
        result = subprocess.run(['git', '-C', REPO_PATH, 'diff', '--cached', '--stat'],
                                capture_output=True, text=True)
        if not result.stdout.strip():
            print("  ○ Sin cambios nuevos en git — sync ya actualizado")
            return
        subprocess.run(['git', '-C', REPO_PATH, 'commit', '-m', msg],
                       check=True, capture_output=True)
        subprocess.run(['git', '-C', REPO_PATH, 'push'],
                       check=True, capture_output=True)
        print(f"  ✓ Git push exitoso: \"{msg}\"")
    except subprocess.CalledProcessError as e:
        print(f"  ✗ Git falló: {e.stderr.decode() if e.stderr else str(e)}")
        print("  → Los archivos están en disco. Puedes hacer el push manualmente:")
        print(f"    git -C {REPO_PATH} add creditek/data/ creditek/assets/imagenes/ && git commit -m '{msg}' && git push")

# === MAIN ===
def main():
    print("\n" + "="*70)
    print("  CREDITEK · CATÁLOGO ENRIQUECIDO v2 (Gemini + Incremental)")
    print("="*70 + "\n")

    env = cargar_env()
    gemini_key = env.get('GEMINI_API_KEY', '')
    if not gemini_key:
        print("✗ GEMINI_API_KEY no encontrada en .env")
        sys.exit(1)

    print("→ Autenticando con Google Drive...")
    svc = autenticar_drive()
    print("✓ Conectado a Drive\n")

    print("→ Preparando subcarpetas en EQUIPOS APP DE VENTAS...")
    id_specs = buscar_o_crear_carpeta(svc, '_SPECS', CARPETA_BASE)
    id_imgs  = buscar_o_crear_carpeta(svc, '_IMAGENES', CARPETA_BASE)
    print(f"✓ _SPECS: {id_specs}")
    print(f"✓ _IMAGENES: {id_imgs}\n")

    # Cargar índice existente (modo incremental si existe, masivo si no)
    print("→ Verificando índice existente en Drive...")
    indice_existente = cargar_indice_existente(svc)
    modo = 'INCREMENTAL' if indice_existente else 'MASIVO'
    print(f"✓ Modo: {modo} ({len(indice_existente)} referencias ya procesadas)\n")

    print("→ Descargando Excel maestro de Drive...")
    buf = descargar_archivo_drive(svc, EXCEL_FILE_ID)
    wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    if 'Comparativo' not in wb.sheetnames:
        print("✗ Hoja 'Comparativo' no encontrada")
        return
    ws = wb['Comparativo']

    referencias = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            prov = row[19]
            ref  = row[20]
            pW   = row[22]
            if not ref or not pW:
                continue
            try:
                pW = float(pW)
            except Exception:
                continue
            if pW <= 0:
                continue
            referencias.append({
                'referencia': str(ref).strip(),
                'proveedor':  str(prov).strip() if prov else '',
                'precio_venta': int(round(pW))
            })
        except Exception:
            pass

    vistos = set()
    unicas = []
    for r in referencias:
        s = slug(r['referencia'])
        if s not in vistos:
            unicas.append(r)
            vistos.add(s)

    # Determinar cuáles procesar en este run
    a_procesar = []
    a_saltar   = []
    for r in unicas:
        s = slug(r['referencia'])
        if s not in indice_existente:
            a_procesar.append(('nuevo', r))
        else:
            existente = indice_existente[s]
            marca = detectar_marca(r['referencia'])
            if not existente['tiene_specs']:
                # Sin specs aún (primera ejecución falló o ref nueva) → procesar completo
                a_procesar.append(('nuevo', r))
            elif not existente['tiene_imagen'] and es_marca_reconocida(marca):
                # Tiene specs pero sin imagen → solo reintentar imagen
                a_procesar.append(('reintentar_imagen', r))
            else:
                a_saltar.append(r)

    print(f"✓ {len(unicas)} referencias únicas en Excel")
    print(f"  → A procesar: {len(a_procesar)} ({sum(1 for t,_ in a_procesar if t=='nuevo')} nuevas, "
          f"{sum(1 for t,_ in a_procesar if t=='reintentar_imagen')} reintentos de imagen)")
    print(f"  → A saltar:   {len(a_saltar)}")
    print("="*70 + "\n")

    if not a_procesar:
        print("✓ Catálogo al día — nada que procesar\n")
        sincronizar_github_pages(svc, id_specs, id_imgs, indice_existente)
        return

    indice_actualizado = dict(indice_existente)
    hoy = datetime.now().strftime('%Y-%m-%d')
    total = len(a_procesar)

    for i, (tipo, r) in enumerate(a_procesar, 1):
        ref   = r['referencia']
        marca = detectar_marca(ref)
        slug_nombre = slug(ref)
        json_name   = f"{slug_nombre}.json"
        img_name    = f"{slug_nombre}.jpg"

        print(f"[{i:3}/{total}] {ref[:55]:55} | {marca:10} | {tipo}", end=' ', flush=True)

        tiene_imagen = False
        notas        = ''

        if tipo == 'reintentar_imagen':
            # Reusar specs ya guardadas, solo reintentar imagen
            data_gemini = consultar_gemini(ref, gemini_key)
            time.sleep(GEMINI_DELAY)
            if data_gemini and data_gemini.get('imagen_url'):
                img_bytes = descargar_imagen(data_gemini['imagen_url'])
                if img_bytes:
                    subir_archivo(svc, img_name, img_bytes, 'image/jpeg', id_imgs)
                    tiene_imagen = True
            if not tiene_imagen:
                img_bytes = buscar_imagen_wikipedia(ref)
                if img_bytes:
                    subir_archivo(svc, img_name, img_bytes, 'image/jpeg', id_imgs)
                    tiene_imagen = True
            print(f"→ {'✓ img' if tiene_imagen else '✗ sin img'}")
            # Actualizar índice
            row_ant = indice_existente.get(slug_nombre, {})
            indice_actualizado[slug_nombre] = {
                **row_ant,
                'tiene_imagen': tiene_imagen,
                'fecha': hoy,
                'notas': notas,
            }
            continue

        # Tipo 'nuevo': procesar completo
        ficha = {
            'referencia': ref,
            'marca': marca,
            'pantalla': '',
            'ram_almacenamiento': '',
            'camara': '',
            'bateria': '',
            'red': '',
            'sistema': '',
            'fuente': '',
            'confianza': '',
            'fecha_extraccion': hoy
        }

        if es_marca_reconocida(marca):
            data_gemini = consultar_gemini(ref, gemini_key)
            time.sleep(GEMINI_DELAY)
            if data_gemini:
                for k in ['pantalla', 'ram_almacenamiento', 'camara', 'bateria', 'red', 'sistema', 'confianza']:
                    ficha[k] = data_gemini.get(k, '')
                ficha['fuente'] = 'Gemini API'
                # Intentar imagen
                img_bytes = descargar_imagen(data_gemini.get('imagen_url'))
                if img_bytes:
                    subir_archivo(svc, img_name, img_bytes, 'image/jpeg', id_imgs)
                    tiene_imagen = True
                if not tiene_imagen:
                    img_bytes = buscar_imagen_wikipedia(ref)
                    if img_bytes:
                        subir_archivo(svc, img_name, img_bytes, 'image/jpeg', id_imgs)
                        tiene_imagen = True
                print(f"→ ✓ specs{'+img' if tiene_imagen else ''} [{ficha['confianza']}]")
            else:
                notas = 'Error en Gemini API'
                print(f"→ ✗ error Gemini")
        else:
            # Marca genérica: specs básicas vía Gemini, sin imagen
            data_gemini = consultar_gemini(ref, gemini_key)
            time.sleep(GEMINI_DELAY)
            if data_gemini:
                for k in ['pantalla', 'ram_almacenamiento', 'camara', 'bateria', 'red', 'sistema', 'confianza']:
                    ficha[k] = data_gemini.get(k, '')
                ficha['fuente'] = 'Gemini API (genérica)'
                print(f"→ ○ specs básicas [{ficha['confianza']}]")
            else:
                notas = 'Marca genérica, error Gemini'
                print(f"→ ✗ error Gemini")

        # Subir JSON de specs
        json_bytes = json.dumps(ficha, indent=2, ensure_ascii=False).encode('utf-8')
        subir_archivo(svc, json_name, json_bytes, 'application/json', id_specs)

        indice_actualizado[slug_nombre] = {
            'referencia': ref,
            'marca': marca,
            'tiene_imagen': tiene_imagen,
            'tiene_specs': bool(any(ficha.get(k, '') for k in
                                   ['pantalla', 'ram_almacenamiento', 'camara', 'bateria', 'red', 'sistema'])),
            'fuente': ficha['fuente'][:80],
            'fecha': hoy,
            'notas': notas,
        }

    # Actualizar INDICE.xlsx en Drive
    print("\n→ Actualizando INDICE.xlsx en Drive...")
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = 'INDICE'
    ws2.append(['Referencia', 'Marca', 'Tiene imagen', 'Tiene specs', 'Fuente', 'Fecha', 'Notas'])
    for row in indice_actualizado.values():
        ws2.append([
            row['referencia'], row['marca'],
            '✓' if row['tiene_imagen'] else '✗',
            '✓' if row['tiene_specs'] else '✗',
            row['fuente'], row['fecha'], row['notas']
        ])
    for col in ws2.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
    buf2 = io.BytesIO()
    wb2.save(buf2)
    subir_archivo(svc, 'INDICE.xlsx', buf2.getvalue(),
                  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                  CARPETA_BASE)

    # Resumen
    total_idx  = len(indice_actualizado)
    con_img    = sum(1 for v in indice_actualizado.values() if v['tiene_imagen'])
    con_specs  = sum(1 for v in indice_actualizado.values() if v['tiene_specs'])
    sin_nada   = sum(1 for v in indice_actualizado.values() if not v['tiene_imagen'] and not v['tiene_specs'])

    print("\n" + "="*70)
    print("  RESUMEN")
    print("="*70)
    print(f"  Total en índice:            {total_idx}")
    print(f"  Con imagen + specs:         {con_img}")
    print(f"  Solo specs (sin imagen):    {con_specs - con_img}")
    print(f"  Solo info básica:           {sin_nada}")
    print(f"  Procesadas en este run:     {len(a_procesar)}")
    print("="*70)

    # Sync a GitHub Pages
    sincronizar_github_pages(svc, id_specs, id_imgs, indice_actualizado)


if __name__ == '__main__':
    main()
